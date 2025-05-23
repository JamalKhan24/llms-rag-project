import re
import json
from openpyxl import load_workbook

xlsx_file = "data_preprocessing/data/NUST Bank-Product-Knowledge.xlsx"
wb = load_workbook(xlsx_file, data_only=True)
sheets = wb.sheetnames

qa_data = {}

question_pattern = re.compile(
    r"(?i)\b(what|how|is|does|can|should|are|do|where|when|why|who|which|whom|whose|could|would|will|shall|has|have|had|may|might|must).*[.?\n]")


def is_question(cell):
    if cell is None or str(cell).strip() == "":
        return False
    text = str(cell).strip()
    return text.endswith('?') or (question_pattern.match(text) and text.endswith('.'))


def clean_line(text):
    return re.sub(r"\s{2,}", " ", str(text)).strip()


def clean_table_answer(answer_list):
    clean = []
    skip_next = 0
    for line in answer_list:
        line = str(line).strip()
        if "Profit Payment" in line and "Profit Rate" in line:
            skip_next = 3
            continue
        if skip_next > 0:
            skip_next -= 1
            continue
        if line.lower() in ["main", "home", "button", "click here"]:
            continue
        if line:
            clean.append(clean_line(line))
    return clean


for sheet_name in sheets[2:]:
    sheet = wb[sheet_name]

    # Detect actual starting column (0 or 1)
    sample_row = next(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
    start_col = 0 if sample_row[0] and str(sample_row[0]).strip() else 1

    # Get title from correct cell
    title = sheet.cell(1, start_col + 1).value  # openpyxl is 1-indexed
    qa_list = []

    question = None
    answer_lines = []

    for row in sheet.iter_rows(values_only=True):
        row = row[start_col:]
        row = [str(cell).strip() for cell in row if cell is not None and str(cell).strip() != ""]
        if not row:
            continue

        cell_text = row[0]

        if is_question(cell_text):
            if question:
                answer = clean_table_answer(answer_lines)
                if answer:
                    qa_list.append({
                        "question": question,
                        "answer": " ".join(answer)
                    })
            question = cell_text
            answer_lines = row[1:]
        else:
            answer_lines.extend(row)

    if question and answer_lines:
        answer = clean_table_answer(answer_lines)
        if answer:
            qa_list.append({
                "question": question,
                "answer": " ".join(answer)
            })

    qa_data[sheet_name] = {
        "title": clean_line(title) if title else None,
        "qa_pairs": qa_list
    }

with open("data_preprocessing/output/nust_accounts_qa.json", "w", encoding="utf-8") as f:
    json.dump(qa_data, f, indent=2, ensure_ascii=False)

print("QA extraction complete. Titles and questions handled correctly.")
