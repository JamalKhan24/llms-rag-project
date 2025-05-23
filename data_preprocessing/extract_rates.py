import json
from openpyxl import load_workbook

# Load Excel sheet 2: "Rate Sheet July 1 2024"
xlsx_file = "data_preprocessing/data/NUST Bank-Product-Knowledge.xlsx"
sheet_name = "Rate Sheet July 1 2024"
wb = load_workbook(xlsx_file, data_only=True)
ws = wb[sheet_name]

rates_data = {
    "Savings Accounts": {},
    "Term Deposits": {}
}

current_section = None
current_account = None
header = None
start_collecting = False

def is_blank_row(row):
    return all([cell is None or str(cell).strip() == "" for cell in row])

for row in ws.iter_rows(values_only=True):
    row = [str(cell).strip() if cell is not None else "" for cell in row]
    if is_blank_row(row):
        continue

    # Detect Sections
    if "Savings Account Profit Rates" in row[0]:
        current_section = "Savings Accounts"
        continue
    if "Term Deposit Profit Rates" in row[0]:
        current_section = "Term Deposits"
        continue

    # Detect Account name (bold-like headers)
    if row[0] and not any(keyword in row[0].lower() for keyword in ["profit", "tenor", "rate", "payout"]) and current_section:
        current_account = row[0]
        header = None
        start_collecting = False
        continue

    # Detect Header
    if current_account and any("profit" in cell.lower() or "tenor" in cell.lower() for cell in row):
        header = row
        start_collecting = True
        continue

    # Parse data row under header
    if start_collecting and header:
        if current_section == "Savings Accounts":
            entry = {}
            for i in range(min(len(header), len(row))):
                key = header[i]
                val = row[i]
                if key and val:
                    entry[key] = val
            if current_account not in rates_data[current_section]:
                rates_data[current_section][current_account] = []
            rates_data[current_section][current_account].append(entry)

        elif current_section == "Term Deposits":
            entry = {}
            for i in range(min(len(header), len(row))):
                key = header[i]
                val = row[i]
                if key and val:
                    entry[key] = val
            if current_account not in rates_data[current_section]:
                rates_data[current_section][current_account] = []
            rates_data[current_section][current_account].append(entry)

# Final cleanup: standardize keys
for section in rates_data:
    for account, entries in rates_data[section].items():
        for i, item in enumerate(entries):
            clean_entry = {}
            for k, v in item.items():
                k_clean = k.strip()
                if k_clean.lower().startswith("profit payment"):
                    clean_entry["Profit Payment"] = v
                elif k_clean.lower().startswith("profit rate"):
                    clean_entry["Profit Rate"] = v
                elif k_clean.lower().startswith("tenor"):
                    clean_entry["Tenor"] = v
                elif k_clean.lower().startswith("payout"):
                    clean_entry["Payout"] = v
                else:
                    clean_entry[k_clean] = v
            entries[i] = clean_entry

# Save to output
with open("data_preprocessing/output/nust_rate_list.json", "w", encoding="utf-8") as f:
    json.dump(rates_data, f, indent=2, ensure_ascii=False)

print("Rate sheet extracted and saved to output/nust_rate_list.json")
